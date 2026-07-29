#!/usr/bin/env python3
"""
菜品数据 Excel 导入脚本
将同事填好的 菜品数据_v2.xlsx 导回 dish_pool.json。
- 按 dish ID 匹配
- 只更新非空单元格（空单元格不覆盖已有数据）
- 自动备份 dish_pool.json
- 打印变更摘要
"""
import json
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

# ============================================================
# 枚举映射表（中文 → 英文规范值）
# ============================================================
REVERSE_PROTEIN = {
    "鱼": "fish", "虾": "shrimp", "其他海鲜": "other_seafood",
    "牛肉": "beef", "猪肉": "pork", "鸡肉": "chicken", "鸡蛋": "egg",
    "豆制品": "tofu", "鱼子酱": "caviar", "其他": "other", "无": "none",
}
REVERSE_TASTE = {"清淡": "light", "正常": "normal", "浓味": "rich", "辣": "spicy"}
REVERSE_COOKING = {
    "蒸": "steam", "煮": "boil", "炒": "stir_fried", "煎": "pan_fried",
    "炖": "stew", "烧": "braise", "焖": "simmer", "烤": "roast",
    "凉拌": "cold_mix", "白灼": "blanched", "温拌": "warm_tossed", "其他": "other",
}
REVERSE_CARB = {
    "米饭": "rice", "粥": "porridge", "面": "noodle", "包点饺子": "dim_sum",
    "粗粮": "coarse_grain", "粗粮（含薯类）": "coarse_grain", "薯类": "coarse_grain", "其他": "other",
}
REVERSE_COMPONENT = {"蛋白质": "protein", "蔬菜": "vegetable", "碳水": "carb"}
REVERSE_CATEGORY = {
    "蛋白质 / 主菜": "protein_main", "蛋类 / 豆制品": "egg_tofu",
    "蔬菜 / 菌菇": "vegetable_mushroom", "汤 / 羹": "soup",
    "主食 / 碳水": "staple_carb", "冷菜 / 凉拌": "cold_dish",
    "一餐型料理": "one_pot_meal", "水果 / 加餐 / 下午茶": "fruit_snack",
}


def cn_to_en_list(cn_str, reverse_map):
    """中文逗号分隔 → 英文列表"""
    if not cn_str or not str(cn_str).strip():
        return []
    parts = [p.strip() for p in str(cn_str).replace("，", ",").split(",")]
    result = []
    for p in parts:
        if p in reverse_map:
            result.append(reverse_map[p])
        elif p:
            result.append(p)
    return result


def cell_str(cell_val):
    """安全取单元格字符串值"""
    if cell_val is None:
        return ""
    return str(cell_val).strip()


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(base, "菜品数据_v2.xlsx")
    json_path = os.path.join(base, "dish_pool.json")

    if not os.path.exists(xlsx_path):
        print(f"❌ 找不到 Excel 文件: {xlsx_path}")
        return

    # 1. 备份
    bak_path = os.path.join(base, f"dish_pool.json.bak.import.{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    shutil.copy2(json_path, bak_path)
    print(f"✅ 已备份: {bak_path}")

    # 2. 读取 Excel
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["菜品数据"]

    # 3. 读取现有 dish_pool.json
    with open(json_path, "r", encoding="utf-8") as f:
        pool = json.load(f)

    # 建立 id → dish 索引
    dish_map = {d["id"]: d for d in pool["dishes"]}

    # 列映射（1-based）
    # A=1 ID, B=2 中文名, C=3 英文名, D=4 分类, E=5 早餐, F=6 午餐, G=7 晚餐,
    # H=8 家宴, I=9 蛋白质, J=10 蔬菜, K=11 口味, L=12 烹饪, M=13 主食类型,
    # N=14 一餐型组成, O=15 可改温热, P=16 自定义标签, Q=17 需审核, R=18 原有食材
    changes_log = []
    updated_count = 0

    for row in ws.iter_rows(min_row=2, max_col=18):
        dish_id = cell_str(row[0].value)
        if not dish_id or not dish_id.startswith("dish_"):
            continue
        if dish_id not in dish_map:
            continue

        d = dish_map[dish_id]
        changed = False

        # 分类
        cat_cn = cell_str(row[3].value)
        if cat_cn and cat_cn in REVERSE_CATEGORY and REVERSE_CATEGORY[cat_cn] != d["category_id"]:
            d["category_id"] = REVERSE_CATEGORY[cat_cn]
            changed = True

        # 餐别标签
        old_tags = set(d.get("meal_tags", []))
        new_tags = set()
        if cell_str(row[4].value) == "✓":
            new_tags.add("breakfast")
        if cell_str(row[5].value) == "✓":
            new_tags.add("lunch")
        if cell_str(row[6].value) == "✓":
            new_tags.add("dinner")
        if new_tags and new_tags != old_tags:
            d["meal_tags"] = sorted(new_tags)
            changed = True

        # 家宴
        banquet_val = cell_str(row[7].value) == "✓"
        if banquet_val != d.get("banquet", False):
            d["banquet"] = banquet_val
            changed = True

        # 蛋白质
        protein_str = cell_str(row[8].value)
        if protein_str:
            new_proteins = cn_to_en_list(protein_str, REVERSE_PROTEIN)
            if new_proteins and new_proteins != d.get("protein_types", []):
                d["protein_types"] = new_proteins
                changed = True

        # 蔬菜
        veg_str = cell_str(row[9].value)
        if veg_str:
            new_vegs = [v.strip() for v in veg_str.replace("，", ",").split(",") if v.strip()]
            if new_vegs and new_vegs != d.get("vegetables", []):
                d["vegetables"] = new_vegs
                d["vegetable_count"] = len(new_vegs)
                changed = True

        # 口味
        taste_cn = cell_str(row[10].value)
        if taste_cn and taste_cn in REVERSE_TASTE and REVERSE_TASTE[taste_cn] != d.get("taste"):
            d["taste"] = REVERSE_TASTE[taste_cn]
            changed = True

        # 烹饪方式
        cooking_str = cell_str(row[11].value)
        if cooking_str:
            new_cookings = cn_to_en_list(cooking_str, REVERSE_COOKING)
            if new_cookings and new_cookings != d.get("cooking_methods", []):
                d["cooking_methods"] = new_cookings
                changed = True

        # 主食类型
        carb_cn = cell_str(row[12].value)
        if carb_cn and carb_cn in REVERSE_CARB:
            new_carb = REVERSE_CARB[carb_cn]
            if new_carb != d.get("carb_type"):
                d["carb_type"] = new_carb
                changed = True

        # 一餐型组成
        comp_str = cell_str(row[13].value)
        if comp_str:
            new_comps = cn_to_en_list(comp_str, REVERSE_COMPONENT)
            if new_comps and new_comps != d.get("meal_components", []):
                d["meal_components"] = new_comps
                changed = True

        # 可改温热
        warm_val = cell_str(row[14].value) == "✓"
        if warm_val != d.get("can_serve_warm", False):
            d["can_serve_warm"] = warm_val
            changed = True

        # 自定义标签
        tag_str = cell_str(row[15].value)
        if tag_str:
            new_tags = [t.strip() for t in tag_str.replace("，", ",").split(",") if t.strip()]
            if new_tags and new_tags != d.get("custom_tags", []):
                d["custom_tags"] = new_tags
                changed = True

        # 需审核
        review_val = cell_str(row[16].value) == "✓"
        if review_val != d.get("needs_review", False):
            d["needs_review"] = review_val
            changed = True

        if changed:
            updated_count += 1
            changes_log.append(dish_id)

    # 4. 保存
    pool["meta"]["last_updated"] = datetime.now().strftime("%Y-%m-%d")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(pool, f, ensure_ascii=False, indent=2)

    # 5. 打印摘要
    print(f"\n✅ 导入完成")
    print(f"   总菜品: {len(pool['dishes'])} 道")
    print(f"   有变更: {updated_count} 道")
    print(f"   无变更: {len(pool['dishes']) - updated_count} 道")
    if changes_log:
        print(f"   变更菜品 ID: {', '.join(changes_log[:20])}")
        if len(changes_log) > 20:
            print(f"   ... 及其他 {len(changes_log)-20} 道")
    print(f"   备份文件: {bak_path}")
    print(f"\n💡 提示: 刷新 photo_manager 页面即可看到新数据")


if __name__ == "__main__":
    main()
