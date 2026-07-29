#!/usr/bin/env python3
"""
菜品数据 Excel 导出脚本
从 dish_pool.json 导出为 菜品数据_v2.xlsx，供多人协作填写。
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# 枚举映射表（英文 → 中文）
# ============================================================
PROTEIN_MAP = {
    "fish": "鱼", "shrimp": "虾", "other_seafood": "其他海鲜", "seafood": "其他海鲜",
    "beef": "牛肉", "pork": "猪肉", "chicken": "鸡肉", "egg": "鸡蛋",
    "tofu": "豆制品", "caviar": "鱼子酱", "other": "其他", "none": "无",
}

TASTE_MAP = {
    "light": "清淡", "normal": "正常", "rich": "浓味", "spicy": "辣",
    "unknown": "", "": "",
}

COOKING_MAP = {
    "steam": "蒸", "steamed": "蒸", "boil": "煮", "boiled": "煮",
    "stir_fried": "炒", "stir_fry": "炒", "pan_fried": "煎", "pan_fry": "煎",
    "stewed": "炖", "stew": "炖", "braise": "烧", "simmer": "焖",
    "roast": "烤", "roasted": "烤", "cold_mix": "凉拌", "cold_mixed": "凉拌",
    "blanched": "白灼", "warm_tossed": "温拌", "other": "其他",
}

CARB_MAP = {
    "rice": "米饭", "porridge": "粥", "noodle": "面", "dim_sum": "包点饺子",
    "coarse_grain": "粗粮（含薯类）", "other": "其他",
}

COMPONENT_MAP = {
    "protein": "蛋白质", "vegetable": "蔬菜", "carb": "碳水",
}

CATEGORY_MAP = {
    "protein_main": "蛋白质 / 主菜", "egg_tofu": "蛋类 / 豆制品",
    "vegetable_mushroom": "蔬菜 / 菌菇", "soup": "汤 / 羹",
    "staple_carb": "主食 / 碳水", "cold_dish": "冷菜 / 凉拌",
    "one_pot_meal": "一餐型料理", "fruit_snack": "水果 / 加餐 / 下午茶",
}

# 反向映射（中文 → 英文规范值）— 导入时用
REVERSE_PROTEIN = {
    "鱼": "fish", "虾": "shrimp", "其他海鲜": "other_seafood",
    "牛肉": "beef", "猪肉": "pork", "鸡肉": "chicken", "鸡蛋": "egg",
    "豆制品": "tofu", "鱼子酱": "caviar", "其他": "other", "无": "none",
}
REVERSE_TASTE = {"清淡": "light", "正常": "normal", "浓味": "rich", "辣": "spicy"}
REVERSE_COOKING = {
    "蒸": "steam", "煮": "boil", "炒": "stir_fried", "煎": "pan_fried",
    "炖": "stew", "烧": "braise", "焖": "simmer", "烤": "roast",
    "凉拌": "cold_mix", "白灼": "blanched", "温拌": "warm_tossed", "其他": "other",
}
REVERSE_CARB = {
    "米饭": "rice", "粥": "porridge", "面": "noodle", "包点饺子": "dim_sum",
    "粗粮": "coarse_grain", "粗粮（含薯类）": "coarse_grain", "薯类": "coarse_grain", "其他": "other",
}
REVERSE_COMPONENT = {"蛋白质": "protein", "蔬菜": "vegetable", "碳水": "carb"}
REVERSE_CATEGORY = {v: k for k, v in CATEGORY_MAP.items()}

# ============================================================
# 样式定义
# ============================================================
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
READONLY_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
NEEDS_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅黄=待填
FILLED_OK = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 浅绿=已填
BANQUET_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 家宴标记
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def en_to_cn_list(en_list, mapping):
    """将英文枚举列表转为中文逗号分隔字符串"""
    return ", ".join(mapping.get(v, v) for v in en_list) if en_list else ""


def cn_to_en_list(cn_str, reverse_map):
    """将中文逗号分隔字符串转为英文枚举列表"""
    if not cn_str or not cn_str.strip():
        return []
    parts = [p.strip() for p in cn_str.replace("，", ",").split(",")]
    result = []
    for p in parts:
        if p in reverse_map:
            result.append(reverse_map[p])
        elif p:
            result.append(p)  # 保留无法映射的原文
    return result


def build_dish_sheet(wb, pool):
    """Sheet 1: 菜品数据"""
    ws = wb.active
    ws.title = "菜品数据"
    dishes = pool["dishes"]

    headers = [
        ("ID", 12), ("中文名", 20), ("英文名", 28), ("分类", 16),
        ("早餐", 6), ("午餐", 6), ("晚餐", 6), ("家宴", 6),
        ("主要蛋白质", 22), ("包含蔬菜", 22), ("口味", 10), ("烹饪方式", 20),
        ("主食类型", 12), ("一餐型组成", 14), ("可改温热", 10),
        ("自定义标签", 22), ("需审核", 8), ("原有食材(参考)", 22),
    ]
    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "D2"  # 冻结 ID+中英文名

    cat_labels = [c["label_cn"] for c in pool["categories"] if c["active"]]
    taste_labels = list(TASTE_MAP.values())[:-2]  # 去掉空值
    carb_labels = list(CARB_MAP.values())
    component_labels = list(COMPONENT_MAP.values())

    for i, d in enumerate(dishes):
        r = i + 2
        cat_cn = CATEGORY_MAP.get(d["category_id"], d["category_id"])
        protein_cn = en_to_cn_list(d.get("protein_types", []), PROTEIN_MAP)
        veg_cn = ", ".join(d.get("vegetables", []))
        taste_cn = TASTE_MAP.get(d.get("taste", ""), d.get("taste", ""))
        cooking_cn = en_to_cn_list(d.get("cooking_methods", []), COOKING_MAP)
        carb_cn = CARB_MAP.get(d.get("carb_type"), "") if d.get("carb_type") else ""
        component_cn = en_to_cn_list(d.get("meal_components", []), COMPONENT_MAP)
        warm_cn = "✓" if d.get("can_serve_warm") else ""
        tags_cn = ", ".join(d.get("custom_tags", []))
        review_cn = "✓" if d.get("needs_review") else ""
        ingredients_cn = ", ".join(d.get("ingredients", []))

        row_data = [
            d["id"], d["name_cn"], d["name_en"], cat_cn,
            "✓" if "breakfast" in d.get("meal_tags", []) else "",
            "✓" if "lunch" in d.get("meal_tags", []) else "",
            "✓" if "dinner" in d.get("meal_tags", []) else "",
            "✓" if d.get("banquet") else "",
            protein_cn, veg_cn, taste_cn, cooking_cn,
            carb_cn, component_cn, warm_cn,
            tags_cn, review_cn, ingredients_cn,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col <= 8 else LEFT

            # 只读列灰色背景
            if col in (1, 2, 3, 18):
                cell.fill = READONLY_FILL
            # 家宴行淡橙
            if d.get("banquet") and col == 8:
                cell.fill = BANQUET_FILL
            # 待填字段浅黄提示
            if col in (9, 10, 12) and not val:
                cell.fill = NEEDS_FILL
            if col == 11 and (not val or val == "unknown"):
                cell.fill = NEEDS_FILL
            # 已填字段浅绿
            if col in (9, 10, 12) and val:
                cell.fill = FILLED_OK

    # 数据验证下拉框
    n = len(dishes) + 1
    # 分类下拉
    dv_cat = DataValidation(type="list", formula1='"' + ",".join(cat_labels) + '"', allow_blank=True)
    ws.add_data_validation(dv_cat)
    dv_cat.add(f"D2:D{n}")
    # 口味下拉
    dv_taste = DataValidation(type="list", formula1='"' + ",".join(taste_labels) + '"', allow_blank=True)
    ws.add_data_validation(dv_taste)
    dv_taste.add(f"K2:K{n}")
    # 主食类型下拉
    dv_carb = DataValidation(type="list", formula1='"' + ",".join(carb_labels) + '"', allow_blank=True)
    ws.add_data_validation(dv_carb)
    dv_carb.add(f"M2:M{n}")
    # 一餐型组成下拉
    dv_comp = DataValidation(type="list", formula1='"' + ",".join(component_labels) + '"', allow_blank=True)
    ws.add_data_validation(dv_comp)
    dv_comp.add(f"N2:N{n}")
    # ✓ 列下拉
    dv_check = DataValidation(type="list", formula1='"✓,"', allow_blank=True)
    ws.add_data_validation(dv_check)
    for col_letter in ["E", "F", "G", "H", "O", "Q"]:
        dv_check.add(f"{col_letter}2:{col_letter}{n}")


def build_instructions_sheet(wb):
    """Sheet 3: 填写说明"""
    ws = wb.create_sheet("填写说明")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30

    title = ws.cell(row=1, column=1, value="菜品数据填写说明")
    title.font = Font(name="Microsoft YaHei", bold=True, size=14, color="4472C4")
    ws.merge_cells("A1:C1")

    sections = [
        ("", ""),
        ("一、填写规则", ""),
        ("规则", "说明"),
        ("灰色背景", "只读字段，请勿修改（ID、中英文名、原有食材）"),
        ("浅黄背景", "待填写 — 请补充内容"),
        ("浅绿背景", "已填写 — 可检查修改"),
        ("✓ 符号", "在对应单元格填 ✓ 表示选中，留空表示不选"),
        ("逗号分隔", "蛋白质/蔬菜/烹饪方式等可多选，用英文逗号分隔"),
        ("", ""),
        ("二、主要蛋白质（可选值）", ""),
        ("中文", "英文（系统内部）"),
        ("鱼", "fish"),
        ("虾", "shrimp"),
        ("其他海鲜", "other_seafood"),
        ("牛肉", "beef"),
        ("猪肉", "pork"),
        ("鸡肉", "chicken"),
        ("鸡蛋", "egg"),
        ("豆制品", "tofu"),
        ("鱼子酱", "caviar"),
        ("其他", "other"),
        ("无", "none"),
        ("", ""),
        ("三、口味（单选）", ""),
        ("清淡", "light — 早餐优先"),
        ("正常", "normal"),
        ("浓味", "rich — 晚餐/家宴"),
        ("辣", "spicy"),
        ("", ""),
        ("四、烹饪方式（可多选）", ""),
        ("蒸", "steam"),
        ("煮", "boil"),
        ("炒", "stir_fried"),
        ("煎", "pan_fried"),
        ("炖", "stew"),
        ("烧", "braise"),
        ("焖", "simmer"),
        ("烤", "roast"),
        ("凉拌", "cold_mix"),
        ("白灼", "blanched"),
        ("温拌", "warm_tossed"),
        ("其他", "other"),
        ("", ""),
        ("五、主食类型（仅主食/碳水类填写）", ""),
        ("米饭", "rice"),
        ("粥", "porridge"),
        ("面", "noodle"),
        ("包点饺子", "dim_sum"),
        ("粗粮（含薯类）", "coarse_grain — 红薯、土豆、芋头、玉米、小米等"),
        ("其他", "other"),
        ("", ""),
        ("六、一餐型组成（仅一餐型料理填写）", ""),
        ("蛋白质", "protein"),
        ("蔬菜", "vegetable"),
        ("碳水", "carb"),
        ("", ""),
        ("七、分类（8个一级分类）", ""),
        ("蛋白质 / 主菜", "protein_main"),
        ("蛋类 / 豆制品", "egg_tofu"),
        ("蔬菜 / 菌菇", "vegetable_mushroom"),
        ("汤 / 羹", "soup"),
        ("主食 / 碳水", "staple_carb"),
        ("冷菜 / 凉拌", "cold_dish"),
        ("一餐型料理", "one_pot_meal"),
        ("水果 / 加餐 / 下午茶", "fruit_snack"),
        ("", ""),
        ("八、协作流程", ""),
        ("步骤1", "本表由系统从 dish_pool.json 自动导出"),
        ("步骤2", "上传到腾讯文档，邀请同事在线协作填写"),
        ("步骤3", "同事重点填写：主要蛋白质、包含蔬菜、烹饪方式"),
        ("步骤4", "填写完成后下载 xlsx 文件"),
        ("步骤5", "在本地运行 import_excel.py 导回系统"),
        ("注意", "导回时只更新非空单元格，不会覆盖同事没填的已有数据"),
        ("注意", "导回前会自动备份 dish_pool.json"),
    ]

    for i, (col_a, col_b) in enumerate(sections, 2):
        ca = ws.cell(row=i, column=1, value=col_a)
        cb = ws.cell(row=i, column=2, value=col_b)
        if col_a and not col_b:
            ca.font = Font(name="Microsoft YaHei", bold=True, size=11, color="4472C4")
        elif col_a == "中文" or col_a == "规则":
            ca.font = Font(name="Microsoft YaHei", bold=True, size=10)
            cb.font = Font(name="Microsoft YaHei", bold=True, size=10)
            ca.fill = HEADER_FILL
            cb.fill = HEADER_FILL
            ca.font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
            cb.font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
        ca.alignment = LEFT
        cb.alignment = LEFT


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(base, "dish_pool.json")
    out_path = os.path.join(base, "菜品数据_v2.xlsx")

    with open(json_path, "r", encoding="utf-8") as f:
        pool = json.load(f)

    wb = Workbook()
    build_dish_sheet(wb, pool)
    build_instructions_sheet(wb)

    wb.save(out_path)
    print(f"✅ 导出成功: {out_path}")
    print(f"   菜品数据: {len(pool['dishes'])} 道")
    print(f"   文件大小: {os.path.getsize(out_path)/1024:.0f} KB")


if __name__ == "__main__":
    main()
